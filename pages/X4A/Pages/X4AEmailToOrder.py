import datetime
import os
import shutil
import random
from CommonUtilities.parse_config import ParseConfigFile
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

from pywinauto import application
from selenium.webdriver.common.by import By

from CommonUtilities.baseSet.BasePage import BasePage
from CommonUtilities.parse_config import ParseConfigFile
from CommonUtilities.readProperties import ReadConfig, parse_config_file
from pynput.keyboard import Key, Controller
from openpyxl import load_workbook
import time

from db.service.X4ABulkOrderDataDbManagementService import X4ABulkOrderDataDbManagementService
from db.service.X4AEmailToOrderDataDbManagementService import X4AEmailToOrderDataDbManagementService

status_file_name = ""


class X4AEmailToOrderPage(BasePage):
    parse_config_json = ParseConfigFile()
    screen_shot_path = ReadConfig.getScreenshotPath()
    system_path = os.getcwd()

    SALES_MENU = (By.XPATH, "//*[@data-testid='sales-MenuItem']")
    EMAIL_TO_ORDER_UPLOAD_OPTION = (By.XPATH, "//*[text()='Email to order']")
    EMAIL_TO_ORDER_UPLOAD_TEXT = (By.XPATH, "//h3[text()='Email Orders']")
    # SEARCH_BUTTON = (By.XPATH, "//button[text()='Search']")
    SEARCH_BUTTON = (By.XPATH, "//*[@id='area2']/div[6]/button[2]")
    RESET_BUTTON = (By.XPATH, "//button[text()='Reset']")
    BACK_SEARCH_BUTTON = (By.XPATH, "//button[text()='Back to search']")
    ETO_FRAME = (By.XPATH, "//*[@title='ETO']")
    HEADER_LIST = (By.XPATH, "//tr[@class='headerRow']/td")
    ETO_ORDER = (By.XPATH, "//*[@class='resultsTable']//tbody//tr[@class='dataRow'][1]")
    DETAIL_HEADER = (By.XPATH, "//*[@id='orderHeaderMain']/div/div[1]/div/div[1]/div[1]/div/div")
    REFERENCE_NO_HEADER = (By.XPATH, "//*[@id='orderHeaderMain']/div/div[2]/div/div[1]/div[1]/div/div")
    SHIP_TO_DETAIL_HEADER = (By.XPATH, "//*[@id='orderHeaderMain']/div/div[2]/div/div[2]/div[1]/div[1]/div")
    SKU_SELECTION_HEADER = (By.XPATH, "//*[@id='orderLinesContainer']/h4")
    ACTION_BUTTON = (By.XPATH,
                     "//*[@id='searchResults']/div[3]/div/div[2]/div/button")  # (Run Part and Price check, Resubmit order, Release to EC Hold, Process order responce)
    ETO_ORDER_STATUS = (By.XPATH,
                        "//*[@id='searchResults']/div[3]/div/div[1]/div")  # (Data extracted, Part & Price Check, Order loaded to ERP, Order released EC hold, Processed)
    DROPDOWN_MENU = (By.XPATH, "//*[@id='menu']")
    DOWNLOAD_EMAIL_OPTION = (By.XPATH, "//div[text()='Download Email']")
    VIEW_EMAIL_BODY_OPTION = (By.XPATH, "//div[text()='View Email Body']")
    VIEW_LOG_OPTION = (By.XPATH, "//div[text()='View Log']")
    ORDER_LOG_POPUP = (By.XPATH, "//span[text()='Order Log']")
    CLOSE_ICON = (By.XPATH, "//*[@data-testid='CloseIcon']")
    CLOSE_BUTTON = (By.XPATH, "//button[text() = 'Close']")
    CANCEL_BUTTON = (By.XPATH, "//button[text() = 'Cancel']")

    SEARCH_DROPDOWN_MENU = (By.XPATH, "//*[@id='area2']/div[2]/select")
    SELECT_CUSTOMER_PO = (By.XPATH, "//*[@value='customerOrderNumber']")
    SELECT_ORDER_STATUS = (By.XPATH, "//*[@value='uiLabel']")
    SEARCH_CUSTOMER_PO = (By.XPATH, "//input[@id='customerOrderNumber']")
    SEARCH_ORDER_STATUS = (By.XPATH, "//input[@id='uiLabel']")
    SEARCH_ETO_ORDER_ID = (By.XPATH, "//input[@id='orderId']")
    SEARCH_IM_ORDER_ID = (By.XPATH, "//input[@id='salesOrderNumber']")

    EMAIL_ORDER_HEADER_LIST = ["Account #", "Country", "Order status", "Customer name", "Customer PO", "Sales order #",
                               "Processed", "Additional Information"]
    USER_DROPDOWN = (By.XPATH, "//*[@data-testid='KeyboardArrowDownIcon']")
    ETO_ERROR_TAB = (By.XPATH, "//*[text()='Email to order']")
    ETO_ERROR_ORDER = (By.XPATH, "(//*[starts-with(@id, 'toolTip')])[1]")
    REFERENCE_NO = (By.XPATH, "//*[text()='Reference numbers']")
    CUSTOMER = (By.XPATH, "//*[text()='Customer']")
    ADDITIONAL_INFO = (By.XPATH, "//*[text()='Additional Information']")
    NOTES = (By.XPATH, "//*[text()='Notes']")
    CUSTOMER_ORDER = (By.XPATH, "(//*[text()='Reference numbers']//parent::div//following-sibling::div//strong)[1]")
    END_USER_ORDER = (By.XPATH, "(//*[text()='Reference numbers']//parent::div//following-sibling::div//strong)[2]")
    SALES_ORDER = (By.XPATH, "(//*[text()='Reference numbers']//parent::div//following-sibling::div//strong)[3]")
    CUSTOMER_NAME = (By.XPATH, "(//*[text()='Customer']//parent::div//following-sibling::div//strong)[1]")
    CUSTOMER_ERP = (By.XPATH, "(//*[text()='Customer']//parent::div//following-sibling::div//strong)[2]")
    ETO_ERROR_ORDER_STATUS = (By.XPATH, "(//*[text()='Customer']//parent::div//following-sibling::div//strong)[3]")
    EMAIL_RECEIVED_TIME = (
    By.XPATH, "(//*[text()='Additional Information']//parent::div//following-sibling::div//strong)[1]")
    SENDER = (By.XPATH, "(//*[text()='Additional Information']//parent::div//following-sibling::div//strong)[2]")
    REQUEST_DELIVERY_DATE = (
    By.XPATH, "(//*[text()='Additional Information']//parent::div//following-sibling::div//strong)[3]")
    BUYER = (By.XPATH, "(//*[text()='Additional Information']//parent::div//following-sibling::div//strong)[4]")
    EMAIL_RESPONSE_TO = (
    By.XPATH, "(//*[text()='Additional Information']//parent::div//following-sibling::div//strong)[5]")
    EMAIL_RESPONSE_BCC = (
    By.XPATH, "(//*[text()='Additional Information']//parent::div//following-sibling::div//strong)[6]")
    ORDER_INFO = (By.XPATH, "(//*[text()='Notes']//parent::div//following-sibling::div//strong)[1]")
    SPECIAL_INSTRUCTION = (By.XPATH, "(//*[text()='Notes']//parent::div//following-sibling::div//strong)[2]")
    SHIPPING_TAB = (By.XPATH, "//*[text()='Shipping']")
    SHIP_TO_INFO = (By.XPATH, "//*[text()='Ship to info']")
    SHIP_TO_INFO_EDIT_ICON = (By.XPATH, "(//*[@data-testid='ModeEditOutlineOutlinedIcon'])[1]")
    SHIP_TO_ID = (By.XPATH, "//*[text()='Ship to ID(suffix):']//strong")
    SHIP_TO_COMPANY_NAME = (By.XPATH, "(//*[text()='Company name:'])[1]//strong")
    SHIP_TO_ADDRESS = (By.XPATH, "((//*[text()='Address:'])[1]//strong)[1]")
    SHIP_TO_CONTACT = (By.XPATH, "(//*[text()='Contact:'])[1]//strong")
    SHIP_TO_PHONE_NO = (By.XPATH, "(//*[text()='Phone number:'])[1]//strong")
    SHIP_TO_EMAIL = (By.XPATH, "(//*[text()='Email:'])[1]//strong")
    END_USER_INFO = (By.XPATH, "//*[text()='End User info']")
    END_USER_INFO_EDIT_ICON = (By.XPATH, "(//*[@data-testid='ModeEditOutlineOutlinedIcon'])[2]")
    END_USER_ID = (By.XPATH, "//*[text()='End User ID(suffix):']//strong")
    END_USER_COMPANY_NAME = (By.XPATH, "(//*[text()='Company name:'])[2]//strong")
    END_USER_ADDRESS = (By.XPATH, "((//*[text()='Address:'])[2]//strong)[1]")
    END_USER_CONTACT = (By.XPATH, "(//*[text()='Contact:'])[2]//strong")
    END_USER_PHONE_NO = (By.XPATH, "(//*[text()='Phone number:'])[2]//strong")
    END_USER_EMAIL = (By.XPATH, "(//*[text()='Email:'])[2]//strong")
    LOGOUT = (By.XPATH, "//*[text()='LogOut']")

    def go_to_email_to_order(self):
        try:
            self.do_click_by_locator(self.SALES_MENU)
            self.logger.info("Clicked on Order in the menu")
            self.do_double_click(self.EMAIL_TO_ORDER_UPLOAD_OPTION)
            self.do_switch_to_required_frame(self.ETO_FRAME)
            self.logger.info("Clicked on Email to order option")
            # self.send_email_to_order()


        except Exception as e:
            self.logger.error('Exception occurred while clicking on Email to order ' + str(e))
            raise e

    def verify_email_to_order_page(self):
        try:
            assert 'Email Orders' in self.get_element_text(
                self.EMAIL_TO_ORDER_UPLOAD_TEXT), "Email Orders Title not present"
            assert 'Search' in self.get_element_text(self.SEARCH_BUTTON), "Search button not present"
            assert 'Reset' in self.get_element_text(self.RESET_BUTTON), "Reset button not present"
            for i in range(1, 8):
                HEADER_LIST = (By.XPATH, "//tr[@class='headerRow']/td[" + str(i) + "]")
                assert self.EMAIL_ORDER_HEADER_LIST[i - 1] in self.get_element_text(HEADER_LIST), \
                    self.EMAIL_ORDER_HEADER_LIST[i - 1] + " not present"
            self.logger.info("verified that Email Orders page")

        except Exception as e:
            self.logger.error('Exception occurred while verifying Email Orders page ' + str(e))
            raise e

    def do_eto_select_order(self):
        try:
            self.do_click_by_locator(self.ETO_ORDER)
            self.logger.info("Clicked on ETO order")

        except Exception as e:
            self.logger.error('Exception occurred while clicking ETO order ' + str(e))
            raise e

    def verify_ETO_order_page_haeder(self):
        try:
            assert 'Details' in self.get_element_text(self.DETAIL_HEADER), "Details header not present"
            assert 'Reference numbers' in self.get_element_text(
                self.REFERENCE_NO_HEADER), "Reference numbers header not present"
            assert 'Ship to details' in self.get_element_text(
                self.SHIP_TO_DETAIL_HEADER), "Ship to details header not present"
            assert 'Customer orders by vendor part number' in self.get_element_text(
                self.SKU_SELECTION_HEADER), "Sku details header not present"
            self.do_click_by_locator(self.DROPDOWN_MENU)
            time.sleep(1)
            assert 'Download Email' in self.get_element_text(
                self.DOWNLOAD_EMAIL_OPTION), "Download Email option not present"
            assert 'View Email Body' in self.get_element_text(
                self.VIEW_EMAIL_BODY_OPTION), "View Email Body option not present"
            assert 'View Log' in self.get_element_text(self.VIEW_LOG_OPTION), "View log option not present"
            self.do_click_by_locator(self.VIEW_LOG_OPTION)
            time.sleep(1)
            assert 'Order Log' in self.get_element_text(self.ORDER_LOG_POPUP), "Order log pop not present"
            assert 'Close' in self.get_element_text(self.CLOSE_BUTTON), "close button not present"
            self.do_click_by_locator(self.CLOSE_BUTTON)
            time.sleep(1)
            self.logger.info("verified that ETO order page headers")
            time.sleep(5)
            self.do_click_by_locator(self.BACK_SEARCH_BUTTON)
            # self.do_click_by_locator(self.CANCEL_BUTTON)

        except Exception as e:
            self.logger.error('Exception occurred while verifying ETO order page header ' + str(e))
            raise e

    def do_eto_select_order_by_status(self):
        try:
            email_to_order_management_srv_obj = X4AEmailToOrderDataDbManagementService()
            db_file_path = ReadConfig.get_db_file_path()
            scenario_detail_list = email_to_order_management_srv_obj.get_scenario_details(db_file_path, '1')

            self.do_click_by_locator(self.SEARCH_CUSTOMER_PO)
            self.do_send_keys(self.SEARCH_CUSTOMER_PO, scenario_detail_list[0][7])

            self.do_click_by_locator(self.SEARCH_DROPDOWN_MENU)
            self.do_click_by_locator(self.SELECT_ORDER_STATUS)
            self.do_click_by_locator(self.SEARCH_ORDER_STATUS)
            self.do_send_keys(self.SEARCH_ORDER_STATUS, 'Completed')

            self.do_click_by_locator(self.SEARCH_BUTTON)
            time.sleep(5)
            self.do_click_by_locator(self.ETO_ORDER)
            self.logger.info("Clicked on ETO order")

        except Exception as e:
            self.logger.error('Exception occurred while clicking ETO order ' + str(e))
            raise e

    def do_eto_select_order_by_PO(self, scenario_number):
        try:
            email_to_order_management_srv_obj = X4AEmailToOrderDataDbManagementService()
            db_file_path = ReadConfig.get_db_file_path()
            scenario_detail_list = email_to_order_management_srv_obj.get_scenario_details(db_file_path, scenario_number)

            self.do_click_by_locator(self.SEARCH_DROPDOWN_MENU)
            self.do_click_by_locator(self.SELECT_CUSTOMER_PO)

            self.do_click_by_locator(self.SEARCH_CUSTOMER_PO)
            self.do_send_keys(self.SEARCH_CUSTOMER_PO, scenario_detail_list[0][7])
            self.do_click_by_locator(self.SEARCH_BUTTON)
            time.sleep(5)
            self.do_click_by_locator(self.ETO_ORDER)
            self.logger.info("Clicked on ETO order")

        except Exception as e:
            self.logger.error('Exception occurred while clicking ETO order ' + str(e))
            raise e

    def verify_ETO_order_status(self):
        try:

            if 'Data Extracted' in self.get_element_text(self.ETO_ORDER_STATUS):
                assert 'Run Part and Price Check' in self.get_element_text(
                    self.ACTION_BUTTON), "Run Part and Price Check button not present"
                self.logger.info("checked Data Extracted")
            elif 'Part and Price Check' in self.get_element_text(self.ETO_ORDER_STATUS):
                assert 'Resubmit order' in self.get_element_text(
                    self.ACTION_BUTTON), "Resubmit order button not present"
                self.logger.info("checked Part and Price Check status")
            elif 'Order Loaded to ERP' in self.get_element_text(self.ETO_ORDER_STATUS):
                assert 'Release Order EC Hold' in self.get_element_text(
                    self.ACTION_BUTTON), "Release EC hold button not present"
                self.logger.info("checked Order Released EC Hold status")
            elif 'Order Released EC Hold' in self.get_element_text(self.ETO_ORDER_STATUS):
                assert 'Process Order Response' in self.get_element_text(
                    self.ACTION_BUTTON), "Process Order Response button not present"
                self.logger.info("checked Order Released EC Hold status")
            elif 'Order Response Sent' in self.get_element_text(self.ETO_ORDER_STATUS):
                self.logger.info("ETO Order status is completed")

            time.sleep(2)
            self.do_click_by_locator(self.BACK_SEARCH_BUTTON)
            self.do_click_by_locator(self.RESET_BUTTON)
        except Exception as e:
            self.logger.error('Exception occurred while verifying ETO order status ' + str(e))
            raise e

    def click_on_eto_tab(self):
        try:
            self.do_click_by_locator(self.ETO_ERROR_TAB)
            return True
        except Exception as e:
            self.logger.error('Exception occurred while clicking on ETO error order tab ' + str(e))
            return False

    def click_on_eto_error_order(self):
        try:
            self.do_click_by_locator(self.ETO_ERROR_ORDER)
            return True
        except Exception as e:
            self.logger.error('Exception occurred while clicking on ETO error order ' + str(e))
            return False

    def validate_eto_error_order_details(self):
        try:
            if self.do_check_visibility(self.REFERENCE_NO) and self.do_check_visibility(
                    self.CUSTOMER) and self.do_check_visibility(self.ADDITIONAL_INFO) and self.do_check_visibility(self.NOTES):
                self.logger.info("The expected sections are displayed on order details page")
                assert self.get_element_text(self.CUSTOMER_ORDER) == "-" or self.get_element_text(
                    self.CUSTOMER_ORDER) != "", "Customer order assertion failed"
                assert self.get_element_text(self.END_USER_ORDER) == "-" or self.get_element_text(
                    self.END_USER_ORDER) != "", "End user order assertion failed"
                assert self.get_element_text(self.SALES_ORDER) == "-" or self.get_element_text(
                    self.SALES_ORDER) != "", "Sales order assertion failed"
                assert self.get_element_text(self.CUSTOMER_NAME) == "-" or self.get_element_text(
                    self.CUSTOMER_NAME) != "", "Customer name assertion failed"
                assert self.get_element_text(self.CUSTOMER_ERP) == "-" or self.get_element_text(
                    self.CUSTOMER_ERP) != "", "Customer ERP assertion failed"
                assert self.get_element_text(self.ETO_ERROR_ORDER_STATUS) == "-" or self.get_element_text(
                    self.ETO_ERROR_ORDER_STATUS) != "", "ETO error order status assertion failed"
                assert self.get_element_text(self.EMAIL_RECEIVED_TIME) == "-" or self.get_element_text(
                    self.EMAIL_RECEIVED_TIME) != "", "Email received time assertion failed"
                assert self.get_element_text(self.SENDER) == "-" or self.get_element_text(
                    self.SENDER) != "", "Sender assertion failed"
                assert self.get_element_text(self.REQUEST_DELIVERY_DATE) == "-" or self.get_element_text(
                    self.REQUEST_DELIVERY_DATE) != "", "Request delivery date assertion failed"
                assert self.get_element_text(self.BUYER) == "-" or self.get_element_text(
                    self.BUYER) != "", "Buyer assertion failed"
                assert self.get_element_text(self.EMAIL_RESPONSE_TO) == "-" or self.get_element_text(
                    self.EMAIL_RESPONSE_TO) != "", "Email response to assertion failed"
                assert self.get_element_text(self.EMAIL_RESPONSE_BCC) == "-" or self.get_element_text(
                    self.EMAIL_RESPONSE_BCC) != "", "Email response BCC assertion failed"
                assert self.get_element_text(self.ORDER_INFO) == "-" or self.get_element_text(
                    self.ORDER_INFO) != "", "Order information assertion failed"
                assert self.get_element_text(self.SPECIAL_INSTRUCTION) == "-" or self.get_element_text(
                    self.SPECIAL_INSTRUCTION) != "", "Special instruction assertion failed"
                return True
            else:
                return False
        except Exception as e:
            self.logger.error('Exception occurred while verifying ETO error order details tab ' + str(e))
            return False

    def click_on_shipping_tab(self):
        try:
            self.do_click_by_locator(self.SHIPPING_TAB)
            return True
        except Exception as e:
            self.logger.error('Exception occurred while clicking on shipping tab ' + str(e))
            return False

    def ship_to_end_user_info_validate(self):
        try:
            if self.do_check_visibility(self.SHIP_TO_INFO) and self.do_check_visibility(
                    self.SHIP_TO_INFO_EDIT_ICON) and self.do_check_visibility(self.END_USER_INFO) and self.do_check_visibility(self.END_USER_INFO_EDIT_ICON):
                self.logger.info("The ship to and end user info is displayed on Shipping tab")
                assert self.get_element_text(self.SHIP_TO_ID) == "-" or self.get_element_text(
                    self.SHIP_TO_ID) != "", "Ship to ID assertion failed"
                assert self.get_element_text(self.SHIP_TO_COMPANY_NAME) == "-" or self.get_element_text(
                    self.SHIP_TO_COMPANY_NAME) != "", "Ship to company name assertion failed"
                assert self.get_element_text(self.SHIP_TO_ADDRESS) == "-" or self.get_element_text(
                    self.SHIP_TO_ADDRESS) != "", "Ship to address assertion failed"
                assert self.get_element_text(self.SHIP_TO_CONTACT) == "-" or self.get_element_text(
                    self.SHIP_TO_CONTACT) != "", "Ship to contact assertion failed"
                assert self.get_element_text(self.SHIP_TO_PHONE_NO) == "-" or self.get_element_text(
                    self.SHIP_TO_PHONE_NO) != "", "Ship to phone number assertion failed"
                assert self.get_element_text(self.SHIP_TO_EMAIL) == "-" or self.get_element_text(
                    self.SHIP_TO_EMAIL) != "", "Ship to email assertion failed"
                assert self.get_element_text(self.END_USER_ID) == "-" or self.get_element_text(
                    self.END_USER_ID) != "", "End user ID assertion failed"
                assert self.get_element_text(self.END_USER_COMPANY_NAME) == "-" or self.get_element_text(
                    self.END_USER_COMPANY_NAME) != "", "End user company name assertion failed"
                assert self.get_element_text(self.END_USER_ADDRESS) == "-" or self.get_element_text(
                    self.END_USER_ADDRESS) != "", "End user address assertion failed"
                assert self.get_element_text(self.END_USER_CONTACT) == "-" or self.get_element_text(
                    self.END_USER_CONTACT) != "", "End user contact assertion failed"
                assert self.get_element_text(self.END_USER_PHONE_NO) == "-" or self.get_element_text(
                    self.END_USER_PHONE_NO) != "", "End user phone number assertion failed"
                assert self.get_element_text(self.END_USER_EMAIL) == "-" or self.get_element_text(
                    self.END_USER_EMAIL) != "", "End user email assertion failed"
                return True
            else:
                return False
        except Exception as e:
            self.logger.error('Exception occurred while verifying ETO error order shipping tab details ' + str(e))
            return False

    def logout_x4a(self):
        try:
            self.do_click_by_locator(self.USER_DROPDOWN)
            self.do_click_by_locator(self.LOGOUT)
            self.logger.info("Logout Successfully")
            return True
        except Exception as e:
            self.logger.error('Exception occurred while Logout X4A ' + str(e))
            return False

    def send_email_for_order(self):
        smtp_server = self.parse_config_json.get_data_from_config_json("x4aSmtpCredentials", "smtp_server", "config.json")
        smtp_port = 587
        smtp_username = self.parse_config_json.get_data_from_config_json("x4aSmtpCredentials", "smtp_username", "config.json")
        smtp_password = self.parse_config_json.get_data_from_config_json("x4aSmtpCredentials", "enc_password", "config.json")
        from_email = self.parse_config_json.get_data_from_config_json("x4aSmtpCredentials", "from_email", "config.json")
        to_email = self.parse_config_json.get_data_from_config_json("x4aSmtpCredentials", "to_email", "config.json")
        subject = self.parse_config_json.get_data_from_config_json("x4aSmtpCredentials", "subject", "config.json")
        body = self.parse_config_json.get_data_from_config_json("x4aSmtpCredentials", "body", "config.json")

        try:
            msg = MIMEMultipart()
            msg['From'] = from_email
            msg['To'] = to_email
            msg['Subject'] = subject
            msg.attach(MIMEText(body))
            ETO_input_file_name = parse_config_file.get_data_from_config_json("inputFile", "ETOInputFileName")
            ETO_input_file_path = '.\\TestData\\' + ETO_input_file_name

            with open(ETO_input_file_path, 'rb') as f:
                attachment = MIMEApplication(f.read(), _subtype='pdf')
                attachment.add_header('Content-Disposition', 'attachment', filename=ETO_input_file_name)
                msg.attach(attachment)

            with smtplib.SMTP(smtp_server, smtp_port) as smtp:
                smtp.starttls()
                smtp.login(smtp_username, smtp_password)
                smtp.send_message(msg)
            time.sleep(60)
        except Exception as e:
            self.logger.error('Exception occurred while Sending Email for order ' + str(e))
            raise e

    # def search_file_name(self, file_name):
    #     self.go_to_email_to_order()
    #
    #     bulk_order_file_name = file_name.split(".xlsx")
    #     self.do_click_by_locator(self.DROPDOWN_MENU)
    #     self.do_click_by_locator(self.FILE_NAME_OPTION)
    #     self.do_click_by_locator(self.TEXT_AREA)
    #     self.do_send_keys(self.TEXT_AREA, bulk_order_file_name[0])
    #     time.sleep(15)
    #
    # def update_input_sheet(self, scenario_no):
    #
    #     bulk_order_file_name = parse_config_file.get_data_from_config_json("inputFile", "bulkOrderInputFileName")
    #     bulk_order_management_srv_obj = X4ABulkOrderDataDbManagementService()
    #     db_file_path = ReadConfig.get_db_file_path()
    #     scenario_detail_list = bulk_order_management_srv_obj.get_scenario_details(db_file_path, scenario_no)
    #     src = bulk_order_file_name
    #     curr_date = '{dt.month}{dt.day}{dt.year}{dt.hour}{dt.minute}{dt.second}'.format(dt=datetime.datetime.today())
    #
    #     bulk_order_file_name = bulk_order_file_name.split(".xlsx")
    #     dst = bulk_order_file_name[0] + '_' + curr_date + '.xlsx'
    #     # Copy File
    #     shutil.copy(src, dst)
    #
    #     # load excel file
    #     workbook = load_workbook(filename=dst)
    #
    #     # open workbook
    #     sheet = workbook.active
    #     if scenario_no != 5:
    #         for i in range(1, 11):
    #             sheet.cell(row=2, column=i).value = scenario_detail_list[0][i + 2]
    #     else:
    #         for k in range(1, 8):
    #             for i in range(1, 11):
    #                 sheet.cell(row=k + 1, column=i).value = scenario_detail_list[k - 1][i + 2]
    #
    #     workbook.save(filename=dst)
    #     dst = dst.split(".\\TestData\\")
    #     return dst[1]
    #
    # def select_file(self, file_name):
    #     try:
    #
    #         bulk_order_file_name = file_name
    #
    #         self.do_click_by_locator(self.BROWSE_BUTTON)
    #         time.sleep(3)
    #         keyboard = Controller()
    #         keyboard.type(self.system_path + "\\TestData\\" + bulk_order_file_name)
    #         keyboard.press(Key.enter)
    #         keyboard.release(Key.enter)
    #         time.sleep(10)
    #
    #
    #     except Exception as e:
    #         self.logger.error('Exception occurred while selecting file ' + str(e))
    #         raise e
    #
    # def get_pagination_first_and_last_page(self):
    #     try:
    #         time.sleep(2)
    #         pages = self.get_all_elements(self.PAGINATION_PAGES)
    #         first_page_number = int(pages[0].text)
    #         last_page_number = int(pages[-1].text)
    #         return first_page_number, last_page_number
    #     except Exception as e:
    #         self.logger.erro("Exception while getting pagination first and last page")
    #         raise e
    #
    # def get_random_page(self, first, last):
    #     try:
    #         if last > 10:
    #             return random.randint(2, 10)
    #         elif last <= 10:
    #             return random.randint(first + 1, last - 1)
    #     except Exception as e:
    #         self.logger.error("Exception while generating random number" + str(e))
    #         raise e
    #
    # def verify_user_name_in_pages(self, user_name):
    #     try:
    #         first_page_number, last_page_number = self.get_pagination_first_and_last_page()
    #         self.logger.info("Verifying user name in page %s", str(first_page_number))
    #         self.go_to_page(first_page_number)
    #         self.verify_user_name_quick_search(user_name)
    #
    #         if first_page_number != last_page_number:
    #             if last_page_number != first_page_number + 1:
    #                 random_page = self.get_random_page(first_page_number, last_page_number)
    #                 self.logger.info("Verifying user name in page %s", str(random_page))
    #                 self.do_select_uploaded_by_option(user_name)
    #                 self.go_to_page(random_page)
    #                 self.verify_user_name_quick_search(user_name)
    #             self.logger.info("Verifying vendor name in page %s", str(last_page_number))
    #             self.do_select_uploaded_by_option(user_name)
    #             self.go_to_page(last_page_number)
    #             time.sleep(4)
    #             self.verify_user_name_quick_search(user_name)
    #         self.logger.info("Successfully verified vendor name")
    #     except Exception as e:
    #         self.logger.error("Exception occurred verifying the vendor name quick search" + str(e))
    #         raise e
    #
    # def verify_user_name_quick_search(self, user_name):
    #     try:
    #         self.logger.info("Verifying the user name in table")
    #         max_rows = self.get_element_text(self.ITEMS_PER_PAGE)
    #         self.logger.info("Max items per page: " + max_rows)
    #         for i in range(int(max_rows)):
    #             if i > 0 and i % 2 == 0:
    #                 table = self.driver.find_element(By.XPATH, self.BULK_ORDER_TABLE)
    #                 self.scroll_down(table)
    #                 time.sleep(2)
    #
    #             self.logger.info("Fetching user name")
    #             user_name_xpath = (By.XPATH, "//div[@class='MuiDataGrid-row'] [@data-id='" + str(
    #                 i) + "']/div[@data-field='createdBy']")
    #             try:
    #                 ui_user_name = self.get_element_text(user_name_xpath)
    #                 self.logger.info("Fetched ui user name :" + str(ui_user_name))
    #             except:
    #                 self.logger.info("There are only " + str(i) + " rows in table")
    #                 break
    #
    #             assert ui_user_name.strip() == user_name.strip(), "user Name mismatched"
    #         self.do_click_by_locator(self.CLOSE_ICON)
    #     except Exception as e:
    #         self.logger.error("Exception occurred verifying the vendor name quick search" + str(e))
    #         raise e
