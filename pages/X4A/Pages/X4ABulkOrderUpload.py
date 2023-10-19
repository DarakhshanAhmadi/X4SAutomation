import datetime
import os
import shutil
import random

from pywinauto import application
from selenium.webdriver.common.by import By

from CommonUtilities.baseSet.BasePage import BasePage
from CommonUtilities.parse_config import ParseConfigFile
from CommonUtilities.readProperties import ReadConfig, parse_config_file
from pynput.keyboard import Key, Controller
from openpyxl import load_workbook
import time

from db.service.X4ABulkOrderDataDbManagementService import X4ABulkOrderDataDbManagementService

status_file_name = ""


class X4ABulkOrderUploadPage(BasePage):
    parse_config_json = ParseConfigFile()
    screen_shot_path = ReadConfig.getScreenshotPath()
    system_path = os.getcwd()

    SALES_MENU = (By.XPATH, "//*[@data-testid='sales-MenuItem']")
    BULK_ORDER_UPLOAD_OPTION = (By.XPATH, "//*[text()='Bulk order upload ']")
    BULK_ORDER_UPLOAD_TEXT = (By.XPATH, "//h3[text()='Bulk order upload']")
    UPLOAD_FILE_BUTTON = (By.XPATH, "//button[text()='Upload file']")
    UPLOAD_FILE_TITLE = (By.XPATH, "//span[text()='Upload file']")
    DOWNLOAD_TEMPLATE_BUTTON = (By.XPATH, "//button[text()='Download template']")
    DISABLED_REVIEW_BUTTON = (By.XPATH, "//button[@disabled and text() = 'Review']")
    DISABLED_PLACE_ORDER_BUTTON = (By.XPATH, "//button[@disabled and text() = 'Place orders']")
    CANCEL_BUTTON = (By.XPATH, "//button[text() = 'Cancel']")
    CLOSE_ICON = (By.XPATH, "//*[@data-testid='CloseIcon']")
    BROWSE_BUTTON = (By.XPATH, '//p[text()="Browse"]')
    ENABLED_CANCEL_BUTTON = (By.XPATH, "//button[not(@disable) and text() = 'Cancel']")
    UPLOAD_FILE_LABEL = (By.XPATH, "//div[@class='MuiBox-root css-zdp87v']")
    CHECKBOX = (By.XPATH, "//div[@data-rowindex='2']/div[@data-field='__check__']")
    SEARCH_BOX = (By.ID, "search")
    SEARCH_BOX_SEARCH_ICON = (By.XPATH, "//*[@data-testid='SearchIcon']")
    FILE_NAME = (By.XPATH, "//div[@class='MuiBox-root css-1oulfbi']")
    DELETE_ICON = (By.XPATH, "//*[@data-testid='DeleteOutlineIcon']")
    REVIEW_BUTTON = (By.XPATH, "//button[text() = 'Review']")
    OK_BUTTON = (By.XPATH, "//button[text() = 'Ok']")
    PLACE_ORDER_BUTTON = (By.XPATH, "//button[text() = 'Place orders']")
    FILE_ERROR_MSG = (By.XPATH, "//div[@class='MuiAlert-message css-acap47-MuiAlert-message']")
    TEMPLATE_ERROR_MSG = (By.XPATH, "//div[@class='MuiAlert-message css-acap47-MuiAlert-message']")
    BACK_BUTTON = (By.XPATH, "//button[text() = 'Back']")
    UPLOAD_FILE_NAME = (By.XPATH, "//span[text()='Upload file']")
    UPLOADED_DATE = (By.XPATH,
                     "//span[@class='MuiTypography-root MuiTypography-bodycopyregular css-6bpapz-MuiTypography-root']/following-sibling::span/following-sibling::span")
    UPLOADED_BY = (
    By.XPATH, "//span[@class='MuiTypography-root MuiTypography-bodycopyregular css-6bpapz-MuiTypography-root']")
    EXCEL_ICON = (By.XPATH, "//*[@data-testid='ExcelIcon']")
    USER_DROPDOWN = (By.XPATH, "//*[@data-testid='KeyboardArrowDownIcon']")
    DROPDOWN_MENU = (By.XPATH, "//*[@data-testid='search-dropdown-SelectOption']")
    STATUS_OPTION = (By.XPATH, "//li[@data-value='documentUploadStatus']")
    UPLOADED_BY_OPTION = (By.XPATH, "//li[@data-value='createdBy']")
    FILE_NAME_OPTION = (By.XPATH, "//li[@data-value='fileName']")
    READY_TO_PLACE_STATUS = (
    By.XPATH, "//span[@class='MuiTypography-root MuiTypography-smalltextregular css-16ksqyg-MuiTypography-root']")
    TEXT_AREA = (By.XPATH, "//*[@data-testid='SearchBar']/div/input")
    ACTION_ICON = (By.XPATH, "//div[@aria-rowindex='4']/div[@data-field='actions']")
    SEARCHED_ACTION_ICON = (By.XPATH, "//div[@aria-rowindex='2']/div[@data-field='actions']")
    LIST_FILE_NAME = (By.XPATH, "//div[@aria-rowindex='2']/div[@data-field='fileName']")
    SEARCHED_FILE_NAME = (By.XPATH, "//div[@aria-rowindex='2']/div[@data-field='fileName']")
    SEARCHED_USER_NAME = (By.XPATH, "//div[@aria-rowindex='2']/div[@data-field='createdBy']")
    SEARCHED_DATE = (By.XPATH, "//div[@aria-rowindex='2']/div[@data-field='createdOn']")
    REVIEW_ACTION_ICON = (By.XPATH, "//*[text()=' Review']")
    VIEW_ACTION_ICON = (By.XPATH, "//*[text()='View']")
    CANCEL_ACTION_ICON = (By.XPATH, "//*[text()=' Cancel']")
    PLACE_ORDER_ACTION_ICON = (By.XPATH, "//*[text()=' Place orders']")
    SEARCHED_FILE_STATUS = (By.XPATH, "//div[@aria-rowindex='2']/div[@data-field='documentUploadStatus']")
    DELETE_ACTION_ICON = (By.XPATH, "//*[text()='Delete']")

    ERROR_ORDER_STATUS = (By.XPATH, "//span[@data-name='Upload details-status-error-found']")
    ORDER_STATUS = (
    By.XPATH, "//span[@class='MuiTypography-root MuiTypography-smalltextregular css-14s1fti-MuiTypography-root']")
    COUNTRY_CODE_ERROR_STATUS = (
    By.XPATH, "//span[@class='MuiTypography-root MuiTypography-smalltextregular css-16ksqyg-MuiTypography-root']")
    EDIT_ICON = (By.XPATH, "//*[@data-testid='EditIcon']")
    EDIT_BUTTON = (By.XPATH, "//button[@data-name='Upload details-status-error-found']")
    OPERATOR_ID = (By.XPATH, "//input[@id='operatorId']")
    COUNTRY_CODE_ID = (By.XPATH, "//input[@id='countryCode']")
    CUSTOMER_BRANCH_AND_NUMBER_ID = (By.XPATH, "//input[@id='customerBranchAndNumber']")
    RESELLER_PO_ID = (By.XPATH, "//input[@id='resellerPo#']")
    CARRIER_CODE_ID = (By.XPATH, "//input[@id='carrierCode']")
    ORDER_TYPE_ID = (By.XPATH, "//input[@id='orderType']")
    HEADER_COMMENT_1_ID = (By.XPATH, "//input[@id='headerComment1']")
    HEADER_COMMENT_2_ID = (By.XPATH, "//input[@id='headerComment2']")
    Ingram_SKU_ID = (By.XPATH, "//input[@id='ingramSku']")
    Qty_ID = (By.XPATH, "//input[@id='qty']")
    DISCARD_BUTTON = (By.XPATH, "//button[text() ='Discard changes']")
    APPLY_BUTTON = (By.XPATH, "//button[text() ='Apply']")
    EXPANDABLE_ICON = (By.XPATH,
                       "//button[@class='MuiButtonBase-root MuiIconButton-root MuiIconButton-sizeMedium css-1hp16lx-MuiButtonBase-root-MuiIconButton-root']")
    DUPLICATE_ERROR_MESSAGE = (By.XPATH, "//div[@class='MuiAlert-message css-acap47-MuiAlert-message']")
    DOWNLOADED_ORDER_LIST_BUTTON = (By.XPATH, "//button[text()='Download order list']")
    SEARCH_MSG = (By.XPATH, "//h2[text()='Add a bulk order!']")
    BULK_ORDER_UPLOAD_LINK = (By.XPATH, "//a[@aria-label='Bulk order upload']")
    ITEMS_PER_PAGE = (By.XPATH,
                      "//div[@class='MuiTablePagination-select MuiSelect-select MuiSelect-standard MuiInputBase-input css-d2iqo8-MuiSelect-select-MuiInputBase-input']")
    BULK_ORDER_TABLE = "//div[@class='MuiDataGrid-virtualScrollerRenderZone css-uw2ren-MuiDataGrid-virtualScrollerRenderZone']"
    ORDER_PAGE_TABLE = "//div[@class='MuiPaper-root MuiPaper-elevation MuiPaper-rounded MuiPaper-elevation4 tss-11quiee-MUIDataTable-paper tss-1x5mjc5-MUIDataTable-root css-z1z8w6-MuiPaper-root']"
    LOGOUT = (By.XPATH, "//*[text()='LogOut']")

    def go_to_bulk_order_upload(self):
        try:
            self.do_click_by_locator(self.SALES_MENU)
            self.logger.info("Clicked on Order in the menu")
            self.do_double_click(self.BULK_ORDER_UPLOAD_OPTION)
            self.logger.info("Clicked on Bulk Order Upload option")

        except Exception as e:
            self.logger.error('Exception occurred while clicking on Bulk Order Upload ' + str(e))
            raise e

    def verify_bulk_order_update_page(self):
        try:
            assert 'Bulk order upload' in self.get_element_text(
                self.BULK_ORDER_UPLOAD_TEXT), "Bulk Order Upload Title not present"
            assert 'Upload file' in self.get_element_text(
                self.UPLOAD_FILE_BUTTON), "Bulk Order Upload Title not present"
            assert 'Download template' in self.get_element_text(
                self.DOWNLOAD_TEMPLATE_BUTTON), "Bulk Order Upload Title not present"
            self.logger.info("verified that bulk order upload page")

        except Exception as e:
            self.logger.error('Exception occurred while verifying bulk order detail page ' + str(e))
            raise e

    def click_upload_file(self):
        try:
            self.do_click_by_locator(self.UPLOAD_FILE_BUTTON)
            self.logger.info("Clicked on upload file button")

        except Exception as e:
            self.logger.error('Exception occurred while clicking upload file button ' + str(e))
            raise e

    def verify_upload_file_popup(self):
        try:
            upload_file_label_msg = 'Drag and Drop file here or Browse files from your computer Supported file type: Excel (.xls, .xlsx)'
            address = self.get_element_text(self.UPLOAD_FILE_LABEL).replace("\n", " ")

            assert 'Review' in self.get_element_text(self.DISABLED_REVIEW_BUTTON), "Disabled Review button not present"
            assert 'Place orders' in self.get_element_text(
                self.DISABLED_PLACE_ORDER_BUTTON), "Disabled Place Order button not present"
            assert 'Cancel' in self.get_element_text(self.CANCEL_BUTTON), "Cancel button not present"
            assert address in upload_file_label_msg, "label message not present"
            self.logger.info("verified that upload file popup")
            time.sleep(5)
            # self.do_click_by_locator(self.CANCEL_BUTTON)

        except Exception as e:
            self.logger.error('Exception occurred while verifying upload file popup is closed ' + str(e))
            raise e

    def do_upload_error_file(self):
        try:
            self.do_click_by_locator(self.BROWSE_BUTTON)
            time.sleep(3)
            keyboard = Controller()
            keyboard.type(self.system_path + '\\TestData\\BulkOrderErrorFile.txt')
            keyboard.press(Key.enter)
            keyboard.release(Key.enter)
            time.sleep(5)

        except Exception as e:
            self.logger.error('Exception occurred while selecting file ' + str(e))
            raise e

    def verify_error_popup(self):
        try:
            file_error_msg = 'File type not supported please try again by uploading an Excel file'
            address = self.get_element_text(self.FILE_ERROR_MSG).replace("\n", " ")

            assert file_error_msg in address, "Bulk Order Upload file error message not present"
            self.logger.info("verified that bulk order upload file error")
            time.sleep(5)

        except Exception as e:
            self.logger.error('Exception occurred while verifying bulk order upload file error ' + str(e))
            raise e

    def do_upload_template_error_file(self):
        try:
            self.do_click_by_locator(self.BROWSE_BUTTON)
            time.sleep(3)
            keyboard = Controller()
            keyboard.type(self.system_path + '\\TestData\\Ingram_Testdata_Template_error_file.xlsx')
            keyboard.press(Key.enter)
            keyboard.release(Key.enter)
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while selecting file ' + str(e))
            raise e

    def verify_template_error_message(self):
        try:

            template_error_msg = 'Template mismatch The provided file contains not supported data by the IM standard ' \
                                 'template, mandatory colums are missing.'
            address = self.get_element_text(self.TEMPLATE_ERROR_MSG).replace("\n", " ")
            assert template_error_msg in address, "Bulk Order Upload file template error message not present"
            assert self.is_present(self.DELETE_ICON), "Delete Icon not found"
            self.logger.info("verified that bulk order upload template file error")

            self.do_click_by_locator(self.DELETE_ICON)
            self.do_click_by_locator(self.CANCEL_BUTTON)
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying bulk order upload file error ' + str(e))
            raise e

    def do_select_file(self, scenario_no):
        try:
            global multiple_order_file_name
            self.go_to_bulk_order_upload()
            multiple_order_file_name = self.update_input_sheet(int(scenario_no))
            self.click_upload_file()
            self.select_file(multiple_order_file_name)

        except Exception as e:
            self.logger.error('Exception occurred while selecting file ' + str(e))
            raise e

    def verify_selected_file_popup(self):
        try:
            upload_file_label_msg = 'Your file is ready to be processed to place orders.'
            address = self.get_element_text(self.UPLOAD_FILE_LABEL).replace("\n", " ")

            assert 'Review' in self.get_element_text(self.REVIEW_BUTTON), "Disabled Review button not present"
            assert 'Place orders' in self.get_element_text(
                self.PLACE_ORDER_BUTTON), "Disabled Place Order button not present"
            assert 'Cancel' in self.get_element_text(self.CANCEL_BUTTON), "Cancel button not present"
            assert self.is_present(self.DELETE_ICON), "Delete Icon not found"
            assert upload_file_label_msg in address, "label message not present"
            self.logger.info("verified that selected file popup")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying selecting file popup is closed ' + str(e))
            raise e

    def do_delete_file(self):
        try:
            self.do_click_by_locator(self.DELETE_ICON)
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while deleting file ' + str(e))
            raise e

    def do_select_review(self):
        try:
            global order_file_name
            order_file_name = self.update_input_sheet(4)
            self.select_file(order_file_name)
            self.do_click_review_button()

        except Exception as e:
            self.logger.error('Exception occurred while selecting review button ' + str(e))
            raise e

    def do_click_review_button(self):
        try:

            self.do_click_by_locator(self.REVIEW_BUTTON)
            self.logger.info("Clicked on review button")
            time.sleep(15)

        except Exception as e:
            self.logger.error('Exception occurred while clicking review button ' + str(e))
            raise e

    def format_date(self, date):
        try:
            self.logger.info("Going to format date : " + str(date))

            date_parts = date.split("/")
            month = int(date_parts[0])
            day = int(date_parts[1])
            year = int(date_parts[2])
            formatted_date = str(month) + "/" + str(day) + "/" + str(year)
            self.logger.info(" Formatted date: " + str(formatted_date))
            return formatted_date
        except Exception as e:
            self.logger.error("Error while formatting the date")
            raise e

    def verify_bulk_order_page(self):
        try:
            bulk_order_file_name = order_file_name.split(".xlsx")
            UPLOAD_FILE_NAME = (By.XPATH, "//span[text()='" + str(bulk_order_file_name[0]) + "']")
            self.logger.info("File Name Updated")
            
            curr_date = '{dt.month}/{dt.day}/{dt.year}'.format(dt=datetime.date.today())
            UPLOADED_DATE = self.format_date(self.get_element_text(self.UPLOADED_DATE))
            self.logger.info("Date updated")
            assert self.is_present(self.EXCEL_ICON), "Excel Icon not found"
            assert bulk_order_file_name[0] in self.get_element_text(UPLOAD_FILE_NAME), "Uploaded file name not present"
            assert curr_date in UPLOADED_DATE, "Uploaded date not current date"
            assert 'Cancel' in self.get_element_text(self.CANCEL_BUTTON), "Cancel button not present"

            self.logger.info("verified that bulk order page")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying bulk order upload page ' + str(e))
            raise e

    def verify_place_order_button(self):
        try:
            #
            assert 'Place orders' in self.get_element_text(self.PLACE_ORDER_BUTTON), "Place Order button not present"
            assert 'Ready to place' in self.get_element_text(self.READY_TO_PLACE_STATUS), "status is not correct"

            self.logger.info("verified that place order button")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying Place Order button ' + str(e))
            raise e

    def do_select_file_name_search(self):
        try:
            self.do_click_by_locator(self.BULK_ORDER_UPLOAD_LINK)
            bulk_order_file_name = order_file_name.split(".xlsx")
            self.do_click_by_locator(self.DROPDOWN_MENU)
            self.do_click_by_locator(self.FILE_NAME_OPTION)
            self.do_click_by_locator(self.TEXT_AREA)
            self.do_send_keys(self.TEXT_AREA, bulk_order_file_name[0])
            time.sleep(10)
            self.logger.info("Opened the list of desired file name")

        except Exception as e:
            self.logger.error('Exception occurred while finding desired file ' + str(e))
            raise e

    def do_click_download_template(self):
        try:
            self.go_to_bulk_order_upload()
            
            self.do_click_by_locator(self.DOWNLOAD_TEMPLATE_BUTTON)
            app = application.Application().connect(title_re="Save As", found_index=0)
            app.SaveAs.Button.click()
            time.sleep(5)
            self.logger.info("Clicked on download template button")

        except Exception as e:
            self.logger.error('Exception occurred while clicking upload file button ' + str(e))
            raise e

    def do_download_multiple_template(self, no_of_files):
        try:
            for i in range(no_of_files):
                self.do_click_by_locator(self.DOWNLOAD_TEMPLATE_BUTTON)
                app = application.Application().connect(title_re="Save", found_index=0)
                app.SaveAs.Button.click()
                time.sleep(1)

            self.logger.info("Clicked on download template button")

        except Exception as e:
            self.logger.error('Exception occurred while clicking upload file button ' + str(e))
            raise e

    def do_select_status(self, status):
        try:
            global status_file_name
            self.go_to_bulk_order_upload()
            # 
            self.do_click_by_locator(self.DROPDOWN_MENU)
            self.do_click_by_locator(self.STATUS_OPTION)
            self.do_click_by_locator(self.TEXT_AREA)
            self.do_send_keys(self.TEXT_AREA, status)
            time.sleep(10)
            status_file_name = self.get_element_text(self.LIST_FILE_NAME)
            self.logger.info("Opened the list of desired status")

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def verify_review_icon(self, status):
        try:
            self.search_file_name(status_file_name)
            assert 'Review' in self.get_element_text(self.REVIEW_ACTION_ICON), "Review icon not present"
            assert status_file_name in self.get_element_text(self.SEARCHED_FILE_NAME), "file name not present"
            assert status in self.get_element_text(self.SEARCHED_FILE_STATUS), "status is not correct"
            self.logger.info("verified that review icon ")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying review icon ' + str(e))
            raise e

    def verify_view_icon(self, status):
        try:
            # 
            self.search_file_name(status_file_name)
            assert 'View' in self.get_element_text(self.VIEW_ACTION_ICON), "View icon not present"
            assert status_file_name in self.get_element_text(self.SEARCHED_FILE_NAME), "file name not present"
            assert status in self.get_element_text(self.SEARCHED_FILE_STATUS), "status is not correct"

            self.logger.info("verified that view icon ")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying view icon ' + str(e))
            raise e

    def do_click_view_icon(self):
        try:

            self.do_click_by_locator(self.VIEW_ACTION_ICON)
            time.sleep(5)
            self.do_click_by_locator(self.DOWNLOADED_ORDER_LIST_BUTTON)
            app = application.Application().connect(title_re="Save", found_index=0)
            app.SaveAs.Button.click()
            time.sleep(5)
            self.logger.info("Clicked on view icon")

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def do_click_review_icon_and_download_order_list(self):
        try:
            self.do_click_by_locator(self.REVIEW_ACTION_ICON)
            time.sleep(5)
            self.do_click_by_locator(self.DOWNLOADED_ORDER_LIST_BUTTON)
            app = application.Application().connect(title_re="Save", found_index=0)
            app.SaveAs.Button.click()
            time.sleep(5)
            self.logger.info("Clicked on review icon")

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def do_click_review_icon(self):
        try:
            self.do_click_by_locator(self.REVIEW_ACTION_ICON)
            time.sleep(5)
            self.logger.info("Clicked on review icon")

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def verify_cancel_icon(self):
        try:
            self.search_file_name(status_file_name)
            assert 'CANCEL' in self.get_element_text(self.CANCEL_ACTION_ICON), "CANCEL icon not present"
            self.logger.info("verified that CANCEL icon ")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying view icon ' + str(e))
            raise e

    def do_click_cancel_icon(self):
        try:
            self.do_click_by_locator(self.CANCEL_ACTION_ICON)
            time.sleep(7)
            self.logger.info("clicked the cancel icon for desired file Name")

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def do_click_place_order_icon(self):
        try:
            self.do_click_by_locator(self.PLACE_ORDER_ACTION_ICON)
            time.sleep(7)
            self.logger.info("clicked the place order icon for desired file Name")

        except Exception as e:
            self.logger.error('Exception occurred while placing order ' + str(e))
            raise e

    def verify_delete_icon(self):
        try:
            self.search_file_name(status_file_name)
            assert 'DELETE' in self.get_element_text(self.DELETE_ACTION_ICON), "Delete icon not present"
            assert 'Upload cancelled' in self.get_element_text(
                self.SEARCHED_FILE_STATUS), "Upload cancelled status not present"
            self.logger.info("verified that delete icon and upload cancelled status ")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying view icon ' + str(e))
            raise e

    def do_click_delete_icon(self):
        try:
            self.do_click_by_locator(self.DELETE_ACTION_ICON)
            time.sleep(7)
            self.logger.info("Opened the list of desired file Name")

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def verify_deleted_record(self):
        try:
            self.search_file_name(status_file_name)
            assert status_file_name not in self.get_element_text(self.SEARCHED_FILE_NAME), "Delete icon not present"
            self.logger.info("verified that delete icon and upload cancelled status ")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying view icon ' + str(e))
            raise e

    def verify_place_orders_icon(self):
        try:
            assert 'PLACE ORDERS' in self.get_element_text(
                self.PLACE_ORDER_ACTION_ICON), "Place orders icon not present"
            self.logger.info("verified that Place orders icon ")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying view icon ' + str(e))
            raise e

    def do_click_place_orders_icon(self):
        try:
            self.do_click_by_locator(self.PLACE_ORDER_ACTION_ICON)
            time.sleep(15)
            self.logger.info("clicked the place order icon for desired file Name")

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def do_click_place_order_button(self):
        try:
            time.sleep(5)
            self.driver.refresh()
            self.do_click_by_locator(self.PLACE_ORDER_BUTTON)
            time.sleep(25)
            self.logger.info("clicked the place order icon for desired file Name")

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def verify_order_placed(self):
        try:
            
            self.search_file_name(order_file_name)
            assert 'Order placed' in self.get_element_text(self.SEARCHED_FILE_STATUS), "Order placed status not present"
            assert 'View' in self.get_element_text(self.VIEW_ACTION_ICON), "View icon not present"
            self.logger.info("verified that status is Order placed ")
            time.sleep(2)

        except Exception as e:
            self.logger.error('Exception occurred while verifying status is Order placed ' + str(e))
            raise e

    def do_select_uploaded_by_option(self, user_name):
        try:
            global status_file_name
            self.driver.refresh()
            self.go_to_bulk_order_upload()
            self.do_click_by_locator(self.DROPDOWN_MENU)
            self.do_click_by_locator(self.UPLOADED_BY_OPTION)
            self.do_click_by_locator(self.TEXT_AREA)
            self.do_send_keys(self.TEXT_AREA, user_name)
            time.sleep(10)
            status_file_name = self.get_element_text(self.LIST_FILE_NAME)
            self.logger.info("Opened the list of desired status")

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def verify_user_name(self, user_name):
        try:
            self.driver.refresh()
            self.search_file_name(status_file_name)
            assert status_file_name in self.get_element_text(self.SEARCHED_FILE_NAME), "file name not present"
            assert user_name in self.get_element_text(self.SEARCHED_USER_NAME), "user name not correct"
            self.logger.info("verified that user name ")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying user name ' + str(e))
            raise e

    def do_upload_duplicate_file(self):
        try:
            global duplicate_order_file_name
            duplicate_order_file_name = self.update_input_sheet(4)
            self.go_to_bulk_order_upload()
            self.click_upload_file()
            self.select_file(duplicate_order_file_name)
            self.logger.info("selected the file to upload ")
            self.do_click_by_locator(self.REVIEW_BUTTON)
            self.logger.info("file review done ")
            time.sleep(5)
            self.go_to_bulk_order_upload()
            self.click_upload_file()
            self.select_file(duplicate_order_file_name)
            self.logger.info("selected the file to upload ")
            self.do_click_by_locator(self.REVIEW_BUTTON)

        except Exception as e:
            self.logger.error('Exception occurred while selecting duplicate file ' + str(e))
            raise e

    def verify_duplicate_file_error_message(self):
        try:
            error_message = duplicate_order_file_name + ' already exists. Please specify unique name.'
            address = self.get_element_text(self.DUPLICATE_ERROR_MESSAGE).replace("\n", " ")

            assert address in error_message, "file name error present"
            self.do_click_by_locator(self.OK_BUTTON)
            self.logger.info("verified duplicate file error message ")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying duplicate file error message ' + str(e))
            raise e

    def do_uploaded_file_with_null_values(self, Scenario):
        try:
            self.go_to_bulk_order_upload()
            file_name = self.update_input_sheet(Scenario)
            self.click_upload_file()
            self.select_file(file_name)
            self.logger.info("selected the file to upload ")
            self.do_click_by_locator(self.REVIEW_BUTTON)
            self.logger.info("file review done ")
            time.sleep(5)

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def verify_error_message(self, Scenario_no):
        bulk_order_management_srv_obj = X4ABulkOrderDataDbManagementService()
        db_file_path = ReadConfig.get_db_file_path()
        scenario_detail_list = bulk_order_management_srv_obj.get_scenario_details(db_file_path, Scenario_no)
        try:
            match Scenario_no:
                case '1':
                    self.logger.info("Verifying Operator_ID value is NULL")
                    no_of_error_found = self.get_element_text(self.ERROR_ORDER_STATUS)
                    no_of_error_found = no_of_error_found.split(" Error found")
                    assert no_of_error_found[0] == Scenario_no, "no of error not matched"
                    self.do_click_by_locator(self.EDIT_ICON)

                    assert self.do_get_attribute(self.OPERATOR_ID, 'value') == '', "Operator Id is not NULL"
                    self.logger.info("verified Operator_ID value is NULL")
                case '2':
                    self.logger.info("verifying Country code error")

                    assert "Error found - countrycode required/invalid" in self.get_element_text(
                        self.COUNTRY_CODE_ERROR_STATUS), "Order status is not Ready to Place"
                    self.do_click_by_locator(self.BULK_ORDER_UPLOAD_LINK)
                    self.logger.info("verified country code error")
                case '3':
                    self.logger.info("Verifying reseller po#,carrier code, ingram sku is null")
                    # 
                    assert 'Error found' in self.get_element_text(self.ERROR_ORDER_STATUS)
                    self.do_click_by_locator(self.EDIT_ICON)

                    assert self.do_get_attribute(self.RESELLER_PO_ID, 'value') == '', "reseller po# is not NULL"
                    assert self.do_get_attribute(self.CARRIER_CODE_ID, 'value') == '', "carrier code is not NULL"
                    self.do_click_by_locator(self.DISCARD_BUTTON)
                    self.do_click_by_locator(self.EXPANDABLE_ICON)
                    self.do_click_by_locator(self.EDIT_BUTTON)
                    assert self.do_get_attribute(self.Ingram_SKU_ID, 'value') == '', "Ingram sku ID is not NULL"
                    self.do_click_by_locator(self.DISCARD_BUTTON)
                    self.logger.info("verified reseller po#,carrier code, ingram sku is null")
                case _:
                    assert scenario_detail_list[0][3] in self.do_get_attribute(self.OPERATOR_ID,
                                                                               'value'), "Operator Id is not present"
                    assert scenario_detail_list[0][4] in self.do_get_attribute(self.COUNTRY_CODE_ID,
                                                                               'value'), "Country code not present"
                    assert scenario_detail_list[0][5] in self.do_get_attribute(self.CUSTOMER_BRANCH_AND_NUMBER_ID,
                                                                               'value'), "customer branch and number is not present"
                    assert scenario_detail_list[0][6] in self.do_get_attribute(self.RESELLER_PO_ID,
                                                                               'value'), "reseller po# is not present"
                    assert scenario_detail_list[0][7] in self.do_get_attribute(self.CARRIER_CODE_ID,
                                                                               'value'), "carrier code is not present"
                    assert scenario_detail_list[0][11] in self.do_get_attribute(self.Ingram_SKU_ID,
                                                                                'value'), "Ingram sku ID is not present"

            self.logger.info("verified that error details ")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying error details ' + str(e))
            raise e

    def do_click_discard_button(self):
        try:
            self.do_click_by_locator(self.DISCARD_BUTTON)
            self.logger.info("Clicked discard button ")
            self.do_click_by_locator(self.CANCEL_BUTTON)
            self.logger.info("cancelled the order")
            time.sleep(5)

        except Exception as e:
            self.logger.error('Exception occurred while cancelling button ' + str(e))
            raise e

    def do_click_apply_button(self):
        try:
            self.do_click_by_locator(self.EDIT_ICON)
            self.do_click_by_locator(self.RESELLER_PO_ID)
            self.do_send_keys(self.RESELLER_PO_ID, 'SOPHOSMSPMMMYYYY')
            self.logger.info("reseller po is updated ")
            self.do_click_by_locator(self.CARRIER_CODE_ID)
            self.do_send_keys(self.CARRIER_CODE_ID, 'OT')
            self.do_click_by_locator(self.APPLY_BUTTON)
            self.logger.info("Carrier code is updated ")
            time.sleep(2)
            self.do_click_by_locator(self.EXPANDABLE_ICON)
            self.do_click_by_locator(self.EDIT_BUTTON)
            self.do_click_by_locator(self.Ingram_SKU_ID)
            self.do_send_keys(self.Ingram_SKU_ID, '485EEB')
            self.do_click_by_locator(self.APPLY_BUTTON)
            self.logger.info("Ingram Sku ID is updated ")
            time.sleep(5)

        except Exception as e:
            self.logger.error('Exception occurred while selecting status ' + str(e))
            raise e

    def verify_order_status(self, status):
        try:
            assert status in self.get_element_text(self.ORDER_STATUS), "Order status is not Ready to Place"
            self.do_click_by_locator(self.CANCEL_BUTTON)
            self.logger.info("cancelled the order")
            self.logger.info("verified that error details ")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying error details ' + str(e))
            raise e

    def verify_bulk_order_upload_page(self):
        try:
            curr_date = '{dt.month}/{dt.day}/{dt.year}'.format(dt=datetime.date.today())
            UPLOADED_DATE = self.format_date(self.get_element_text(self.SEARCHED_DATE))
            searched_file_name = order_file_name.split(".xlsx")
            assert searched_file_name[0] in self.get_element_text(self.SEARCHED_FILE_NAME), "file name not present"
            assert "Shyam Tiwari" in self.get_element_text(self.SEARCHED_USER_NAME), "user name not correct"
            assert 'PLACE ORDERS' in self.get_element_text(self.SEARCHED_ACTION_ICON), "Review icon not present"
            assert 'Ready to place' in self.get_element_text(self.SEARCHED_FILE_STATUS), "status is not correct"
            assert curr_date in UPLOADED_DATE, "Uploaded date not present"
            self.do_select_uploaded_by_option("Shyam Tiwari")
            self.verify_user_name_in_pages("Shyam Tiwari")
            self.logger.info("verified that bulk order upload page")

        except Exception as e:
            self.logger.error('Exception occurred while verifying bulk order detail page ' + str(e))
            raise e

    def verify_multiple_bulk_order_upload_page(self):
        try:
            bulk_order_file_name = multiple_order_file_name.split(".xlsx")
            self.search_file_name(multiple_order_file_name)
            self.do_click_review_icon()
            UPLOAD_FILE_NAME = (By.XPATH, "//span[text()='" + str(bulk_order_file_name[0]) + "']")
            self.logger.info("File Name Updated")
            assert bulk_order_file_name[0] in self.get_element_text(UPLOAD_FILE_NAME), "Uploaded file name not present"
            for i in range(0, 7):
                IM_ORDER_NO = (By.XPATH, "//*[@data-testid='MuiDataTableBodyCell-0-" + str(i) + "']")
                ORDER_STATUS = (By.XPATH, "//*[@data-testid='MuiDataTableBodyCell-1-" + str(i) + "']")
                CONFIRMATION_ID = (By.XPATH, "//*[@data-testid='MuiDataTableBodyCell-3-" + str(i) + "']")
                ORDER_GROUPING = (By.XPATH, "//*[@data-testid='MuiDataTableBodyCell-2-" + str(i) + "']")

                table = self.driver.find_element(By.XPATH, self.ORDER_PAGE_TABLE)
                self.scroll_down(table)
                time.sleep(2)
                self.logger.info("Table scrolled")
                if self.get_element_text(ORDER_GROUPING) == '1' or self.get_element_text(ORDER_GROUPING) == '7':
                    assert "Order placed" in self.get_element_text(ORDER_STATUS), "Order placed status not present"
                    assert self.get_element_text(IM_ORDER_NO) != '', "Order No is NULL"
                    assert self.get_element_text(CONFIRMATION_ID) != '', "Confirmation ID is NULL"
                    self.logger.info("status check for order grouping " + self.get_element_text(ORDER_GROUPING))
                else:
                    assert "Failed" in self.get_element_text(ORDER_STATUS), "Failed status not present"
                    assert self.get_element_text(IM_ORDER_NO) == '', "Order No is not NULL"
                    assert self.get_element_text(CONFIRMATION_ID) != '', "Confirmation ID is NULL"
                    self.logger.info("status check for order grouping " + self.get_element_text(ORDER_GROUPING))

            self.logger.info("verified that multiple bulk order upload page")
            time.sleep(1)

        except Exception as e:
            self.logger.error('Exception occurred while verifying multiple bulk order upload page ' + str(e))
            raise e

    def logout_x4a(self):
        try:
            self.do_click_by_locator(self.USER_DROPDOWN)
            self.do_click_by_locator(self.LOGOUT)
            self.logger.info("Logout Successfully")
            return True
        except Exception as e:
            self.logger.error('Exception occurred while Logout X4A ' + str(e))
            return False

    def search_file_name(self, file_name):
        self.go_to_bulk_order_upload()
        
        bulk_order_file_name = file_name.split(".xlsx")
        self.do_click_by_locator(self.DROPDOWN_MENU)
        self.do_click_by_locator(self.FILE_NAME_OPTION)
        self.do_click_by_locator(self.TEXT_AREA)
        self.do_send_keys(self.TEXT_AREA, bulk_order_file_name[0])
        time.sleep(15)

    def update_input_sheet(self, scenario_no):

        bulk_order_file_name = parse_config_file.get_data_from_config_json("inputFile", "bulkOrderInputFileName")
        bulk_order_management_srv_obj = X4ABulkOrderDataDbManagementService()
        db_file_path = ReadConfig.get_db_file_path()
        scenario_detail_list = bulk_order_management_srv_obj.get_scenario_details(db_file_path, scenario_no)
        src = bulk_order_file_name
        curr_date = '{dt.month}{dt.day}{dt.year}{dt.hour}{dt.minute}{dt.second}'.format(dt=datetime.datetime.today())

        bulk_order_file_name = bulk_order_file_name.split(".xlsx")
        dst = bulk_order_file_name[0] + '_' + curr_date + '.xlsx'
        # Copy File
        shutil.copy(src, dst)

        # load excel file
        workbook = load_workbook(filename=dst)

        # open workbook
        sheet = workbook.active
        if scenario_no != 5:
            for i in range(1, 11):
                sheet.cell(row=2, column=i).value = scenario_detail_list[0][i + 2]
        else:
            for k in range(1, 8):
                for i in range(1, 11):
                    sheet.cell(row=k + 1, column=i).value = scenario_detail_list[k - 1][i + 2]

        workbook.save(filename=dst)
        dst = dst.split(".\\TestData\\")
        return dst[1]

    def select_file(self, file_name):
        try:

            bulk_order_file_name = file_name
            
            self.do_click_by_locator(self.BROWSE_BUTTON)
            time.sleep(3)
            keyboard = Controller()
            keyboard.type(self.system_path + "\\TestData\\" + bulk_order_file_name)
            keyboard.press(Key.enter)
            keyboard.release(Key.enter)
            time.sleep(10)
            

        except Exception as e:
            self.logger.error('Exception occurred while selecting file ' + str(e))
            raise e

    def get_pagination_first_and_last_page(self):
        try:
            time.sleep(2)
            pages = self.get_all_elements(self.PAGINATION_PAGES)
            first_page_number = int(pages[0].text)
            last_page_number = int(pages[-1].text)
            return first_page_number, last_page_number
        except Exception as e:
            self.logger.erro("Exception while getting pagination first and last page")
            raise e

    def get_random_page(self, first, last):
        try:
            if last > 10:
                return random.randint(2, 10)
            elif last <= 10:
                return random.randint(first + 1, last - 1)
        except Exception as e:
            self.logger.error("Exception while generating random number" + str(e))
            raise e

    def verify_user_name_in_pages(self, user_name):
        try:
            first_page_number, last_page_number = self.get_pagination_first_and_last_page()
            self.logger.info("Verifying user name in page %s", str(first_page_number))
            self.go_to_page(first_page_number)
            self.verify_user_name_quick_search(user_name)

            if first_page_number != last_page_number:
                if last_page_number != first_page_number + 1:
                    random_page = self.get_random_page(first_page_number, last_page_number)
                    self.logger.info("Verifying user name in page %s", str(random_page))
                    self.do_select_uploaded_by_option(user_name)
                    self.go_to_page(random_page)
                    self.verify_user_name_quick_search(user_name)
                self.logger.info("Verifying vendor name in page %s", str(last_page_number))
                self.do_select_uploaded_by_option(user_name)
                self.go_to_page(last_page_number)
                time.sleep(4)
                self.verify_user_name_quick_search(user_name)
            self.logger.info("Successfully verified vendor name")
        except Exception as e:
            self.logger.error("Exception occurred verifying the vendor name quick search" + str(e))
            raise e

    def verify_user_name_quick_search(self, user_name):
        try:
            self.logger.info("Verifying the user name in table")
            max_rows = self.get_element_text(self.ITEMS_PER_PAGE)
            self.logger.info("Max items per page: " + max_rows)
            for i in range(int(max_rows)):
                if i > 0 and i % 2 == 0:
                    table = self.driver.find_element(By.XPATH, self.BULK_ORDER_TABLE)
                    self.scroll_down(table)
                    time.sleep(2)

                self.logger.info("Fetching user name")
                user_name_xpath = (By.XPATH, "//div[@class='MuiDataGrid-row'] [@data-id='" + str(
                    i) + "']/div[@data-field='createdBy']")
                try:
                    ui_user_name = self.get_element_text(user_name_xpath)
                    self.logger.info("Fetched ui user name :" + str(ui_user_name))
                except:
                    self.logger.info("There are only " + str(i) + " rows in table")
                    break

                assert ui_user_name.strip() == user_name.strip(), "user Name mismatched"
            self.do_click_by_locator(self.CLOSE_ICON)
        except Exception as e:
            self.logger.error("Exception occurred verifying the vendor name quick search" + str(e))
            raise e
