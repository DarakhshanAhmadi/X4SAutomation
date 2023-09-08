import glob
import os
import time
from pathlib import Path

from openpyxl.reader.excel import load_workbook

from CommonUtilities.logGeneration import LogGenerator
from CommonUtilities.parse_config import ParseConfigFile
from CommonUtilities.readProperties import ReadConfig
from pages.X4A.Facade.BrowserSet import BrowserSettings
from pages.X4A.Pages.X4ALogin import LoginPage
from pages.X4A.Pages.X4ABulkOrderUpload import X4ABulkOrderUploadPage


class ValidateBulkOrderUploadData:
    logger = LogGenerator.logGen()
    parse_config_json = ParseConfigFile()
    screen_shot_path = ReadConfig.getScreenshotPath()
    mandatory_template_fields = ["Operator ID", "Country Code", "Order Type", "Reseller PO#", "Carrier Code", "Ingram SKU", "Qty"]
    template_column = []
    """constructor of the createOrder Page class"""

    def __init__(self, driver):
        self.driver = driver

    def login(self, feature_file_name, screen_shot):
        clear_browser_and_cache = BrowserSettings(self.driver)
        clear_browser_and_cache.do_clear_browser_history_and_cache()
        login = LoginPage(self.driver)

        try:
            environment = self.parse_config_json.get_data_from_config_json("environment", "environment_type",
                                                                           "config.json")
            if environment == 'Stage':
                username = self.parse_config_json.get_data_from_config_json("x4aStageCredentials", "username",
                                                                            "config.json")
                password = self.parse_config_json.get_data_from_config_json("x4aStageCredentials", "enc_password",
                                                                            "config.json")
            else:
                username = self.parse_config_json.get_data_from_config_json("x4aBetaCredentials", "username",
                                                                            "config.json")
                password = self.parse_config_json.get_data_from_config_json("x4aBetaCredentials", "enc_password",
                                                                            "config.json")
            login.do_login_to_x4a(username, password)
            self.logger.info("Successfully logged in to X4A")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_login_successful.png")
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_login_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + "_login_error.png"
            self.logger.error("Login unsuccessful!!")
            self.logger.exception(e)
            raise e

    def filtered_orders_by_feature_file(self, test_data_order, feature_file_name):
        filtered_order_data = test_data_order.loc[(test_data_order.FeatureFileName == feature_file_name)]
        return filtered_order_data

    def click_on_bulk_order_upload_menu(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.go_to_bulk_order_upload()
            self.logger.info("Successfully clicked on error order")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_bulk_order_upload_clicked_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_bulk_order_upload_clicking_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_bulk_order_upload_clicking_error.png"
            self.logger.error("Error while clicking on error order")
            self.logger.exception(e)
            return False

    def verify_bulk_order_upload_page(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_bulk_order_upload_page()
            self.logger.info("Successfully verified bulk order upload page")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_bulk_order_upload_clicked_verified.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_bulk_order_upload_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_bulk_order_upload_verified_error.png"
            self.logger.error("Error while verifying on bulk order upload page")
            self.logger.exception(e)
            return False

    def click_upload_file(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.click_upload_file()
            self.logger.info("Successfully clicked upload file button")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_upload_file_clicked_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_upload_file_clicking_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_upload_file_clicking_error.png"
            self.logger.error("Error while clicking upload file button")
            self.logger.exception(e)
            return False

    def verify_upload_file_popup(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_upload_file_popup()
            self.logger.info("Successfully verified bulk order upload page")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_upload_file_popup_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_upload_file_popup_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_upload_file_popup_verified_error.png"
            self.logger.error("Error while verifying on bulk order upload page %s" + str(e))

    def do_select_file(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            
            x4a_bulk_order_upload.do_select_file()
            self.logger.info("Successfully selected file")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_file_selected_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_file_selected_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_selected_error.png"
            self.logger.error("Error while clicking upload file button")
            self.logger.exception(e)
            return False

    def verify_selected_file_popup(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_selected_file_popup()
            self.logger.info("Successfully verified selected file")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_selected_file_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_selected_file_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_selected_file_verified_error.png"
            self.logger.error("Error while verifying selected file " + str(e))

    def do_upload_error_file(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.do_upload_error_file()
            self.logger.info("Successfully selected file")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_file_selected_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_file_selected_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_selected_error.png"
            self.logger.error("Error while clicking upload file button")
            self.logger.exception(e)
            return False

    def verify_error_popup(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_error_popup()
            self.logger.info("Successfully verified file error")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_selected_file_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_selected_file_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_selected_file_verified_error.png"
            self.logger.error("Error while verifying selected file " + str(e))

    def do_upload_template_error_file(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.do_upload_template_error_file()
            self.logger.info("Successfully uploaded different template file")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_file_selected_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_file_selected_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_selected_error.png"
            self.logger.error("Error while uploading different template file")
            self.logger.exception(e)
            return False

    def verify_template_error_message(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_template_error_message()
            self.logger.info("Successfully verified template error")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_selected_file_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_selected_file_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + "_selected_file_" \
                                                                                                 "verified_error.png"

            self.logger.error("Error while verifying selected file " + str(e))

    def do_delete_file(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.do_delete_file()
            self.logger.info("Successfully deleted file")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_file_deleted_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_file_deleted_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_deleted_error.png"
            self.logger.error("Error while clicking delete icon")
            self.logger.exception(e)
            return False

    def do_select_review(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            
            x4a_bulk_order_upload.do_select_review()
            self.logger.info("Successfully clicked review button")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_file_reviewed_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_file_reviewing_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_reviewing_error.png"
            self.logger.error("Error while clicking review button")
            self.logger.exception(e)
            return False

    def verify_bulk_order_page(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_bulk_order_page()
            self.logger.info("Successfully verified bulk order page")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_bulk_order_page_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_bulk_order_page_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_bulk_order_page_verified_error.png"
            self.logger.error("Error while verifying bulk order page " + str(e))

    def verify_place_order_button(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_place_order_button()
            self.logger.info("Successfully verified place order button")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_place_order_button_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_place_order_button_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_place_order_button_verified_error.png"
            self.logger.error("Error while verifying place order button " + str(e))

    def do_click_download_template(self, feature_file_name, screen_shot, no_of_click):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            
            if no_of_click == "single":
                x4a_bulk_order_upload.do_click_download_template()
            elif no_of_click == "multiple":
                x4a_bulk_order_upload.do_download_multiple_template(5)
            self.logger.info("Successfully clicked Download template button")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_file_downloaded_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_file_downloading_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_downloading_error.png"
            self.logger.error("Error while clicking Download template button")
            self.logger.exception(e)
            return False

    def validate_file_name_and_data(self, feature_file_name, screen_shot):
        try:

            self.logger.info("Fetching latest downloaded file")
            download_dir = str(os.path.join(Path.home(), "Downloads"))
            self.logger.info("Download directory is " + str(download_dir))
            list_of_files = glob.glob(download_dir + '/*.xlsx')
            latest_file = max(list_of_files, key=os.path.getmtime)
            self.logger.info("Recent file in the downloads: " + str(latest_file))

            # load excel file
            workbook = load_workbook(filename=latest_file)

            # open workbook
            sheet_obj = workbook.active
            max_col = sheet_obj.max_column

            # Loop will print all columns name
            for i in range(1, max_col + 1):
                cell_obj = sheet_obj.cell(row=1, column=i)
                self.template_column.append(cell_obj.value)

            for i in self.mandatory_template_fields:
                if i in self.template_column:
                    self.logger.info("Mandatory column " + i + " is present")
                else:
                    self.logger.error("Mandatory column " + i + " is missing")
                time.sleep(1)


            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "downloaded_file_name_verified_successfully.png")
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "verify_downloaded_file_name_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "verify_downloaded_file_name_error.png"
            self.logger.error("Error while verifying the downloaded file name %s" + str(e))
            raise e
        return latest_file

    def validate_multiple_file_name(self, feature_file_name, screen_shot):
        try:
            
            self.logger.info("Fetching latest downloaded file")
            download_dir = str(os.path.join(Path.home(), "Downloads"))
            self.logger.info("Download directory is " + str(download_dir))
            file_list = glob.glob(download_dir + '/Ingram_BulkUpload_Template*.xlsx')
            # latest_file = max(list_of_files, key=os.path.getmtime)
            for i in file_list:
                self.logger.info("Recent file in the downloads: " + str(i))

            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "downloaded_file_name_verified_successfully.png")
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "verify_downloaded_file_name_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "verify_downloaded_file_name_error.png"
            self.logger.error("Error while verifying the downloaded file name %s" + str(e))
            raise e
        return file_list

    def do_select_status(self, feature_file_name, screen_shot, status):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_select_status(status)
            self.logger.info("Successfully selected the status %s", status)
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_status_selected_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_status_selected_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_reviewing_error.png"
            self.logger.error("Error while selecting the status %s", status)
            self.logger.exception(e)
            return False

    def do_select_file_name_search(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_select_file_name_search()
            self.logger.info("Successfully selected the file")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_file_selected_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_file_selected_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_selecting_error.png"
            self.logger.error("Error while selecting the file ")
            self.logger.exception(e)
            return False

    def verify_review_icon(self, feature_file_name, screen_shot, status):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_review_icon(status)
            self.logger.info("Successfully verified review icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_review_icon_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_review_icon_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_review_icon_verified_error.png"
            self.logger.error("Error while verifying review icon " + str(e))

    def verify_view_icon(self, feature_file_name, screen_shot, status):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_view_icon(status)
            self.logger.info("Successfully verified view icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_view_icon_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_view_icon_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_view_icon_verified_error.png"
            self.logger.error("Error while verifying view icon " + str(e))

    def do_click_view_icon(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_click_view_icon()
            self.logger.info("Successfully clicked the view icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_view_icon_clicked_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_view_icon_clicked_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_view_icon_clicked_error.png"
            self.logger.error("Error while clicking the view icon")
            self.logger.exception(e)
            return False

    def do_click_review_icon(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_click_review_icon()
            self.logger.info("Successfully clicked the review icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_review_icon_clicked_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_review_icon_clicked_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_review_icon_clicked_error.png"
            self.logger.error("Error while clicking the review icon")
            self.logger.exception(e)
            return False

    def verify_downloaded_order_list(self, feature_file_name, screen_shot):
        try:
            self.logger.info("Fetching latest downloaded file")
            download_dir = str(os.path.join(Path.home(), "Downloads"))
            self.logger.info("Download directory is " + str(download_dir))
            list_of_files = glob.glob(download_dir + '/*.xlsx')
            latest_file = max(list_of_files, key=os.path.getmtime)
            self.logger.info("Recent file in the downloads: " + str(latest_file))

            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "downloaded_file_name_verified_successfully.png")
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "verify_downloaded_file_name_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "verify_downloaded_file_name_error.png"
            self.logger.error("Error while verifying the downloaded file name %s" + str(e))
            raise e
        return latest_file

    def verify_cancel_icon(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_cancel_icon()
            self.logger.info("Successfully verified CANCEL icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_Cancel_icon_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_Cancel_icon_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_Cancel_icon_verified_error.png"
            self.logger.error("Error while verifying CANCEL icon " + str(e))

    def do_click_cancel_icon(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_click_cancel_icon()
            self.logger.info("Successfully clicked the CANCEL icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_CANCEL_icon_clicked_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_CANCEL_icon_clicked_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_CANCEL_icon_clicked_error.png"
            self.logger.error("Error while clicking the CANCEL icon")
            self.logger.exception(e)
            return False

    def do_click_place_order_icon(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_click_place_order_icon()
            self.logger.info("Successfully clicked the PLACE ORDER icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_place_order_icon_clicked_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_place_order_icon_clicked_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_CANCEL_icon_clicked_error.png"
            self.logger.error("Error while clicking the CANCEL icon")
            self.logger.exception(e)
            return False

    def verify_delete_icon(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_delete_icon()
            self.logger.info("Successfully verified DELETE icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_DELETE_icon_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_DELETE_icon_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_DELETE_icon_verified_error.png"
            self.logger.error("Error while verifying DELETE icon " + str(e))

    def do_click_delete_icon(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_click_delete_icon()
            self.logger.info("Successfully clicked the DELETE icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_DELETE_icon_clicked_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_DELETE_icon_clicked_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_DELETE_icon_clicked_error.png"
            self.logger.error("Error while clicking the DELETE icon")
            self.logger.exception(e)
            return False

    def verify_deleted_record(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_deleted_record()
            self.logger.info("Successfully verified DELETED File not present")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_DELETE_FILE_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_DELETE_FILE_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_DELETE_FILE_verified_error.png"
            self.logger.error("Error while verifying DELETED FILE " + str(e))

    def verify_place_orders_icon(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_place_orders_icon()
            self.logger.info("Successfully verified PLACE ORDERS icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_Place_orders_icon_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_Place_orders_icon_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_Place_orders_icon_verified_error.png"
            self.logger.error("Error while verifying PLACE ORDERS icon " + str(e))

    def do_click_place_orders_icon(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_click_place_orders_icon()
            self.logger.info("Successfully clicked the Place Orders icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_PLACE_ORDERS_icon_clicked_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_PLACE_ORDERS_icon_clicked_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_PLACE_ORDERS_icon_clicked_error.png"
            self.logger.error("Error while clicking the PLACE ORDERS icon")
            self.logger.exception(e)
            return False

    def verify_order_placed(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_order_placed()
            self.logger.info("Successfully verified order placed")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_order_placed_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_order_placed_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_order_placed_verified_error.png"
            self.logger.error("Error while verifying order placed " + str(e))

    def do_select_uploaded_by_option(self, feature_file_name, screen_shot, user_name):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_select_uploaded_by_option(user_name)
            self.logger.info("Successfully selected the status %s", user_name)
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_status_selected_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_status_selected_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_reviewing_error.png"
            self.logger.error("Error while selecting the status %s", user_name)
            self.logger.exception(e)
            return False

    def verify_user_name(self, feature_file_name, screen_shot, user_name):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_user_name(user_name)
            self.logger.info("Successfully verified review icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_review_icon_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_review_icon_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_review_icon_verified_error.png"
            self.logger.error("Error while verifying review icon " + str(e))

    def do_uploaded_file_with_null_values(self, feature_file_name, screen_shot, Scenario):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_uploaded_file_with_null_values(Scenario)
            self.logger.info("Successfully uploaded the file")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_status_selected_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_status_selected_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_reviewing_error.png"
            self.logger.error("Error while selecting the status %s", str(e))
            self.logger.exception(e)
            return False

    def do_click_discard_button(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_click_discard_button()
            self.logger.info("Successfully clicked on discard button")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_status_selected_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_status_selected_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_reviewing_error.png"
            self.logger.error("Error while clicking discard button %s", str(e))
            self.logger.exception(e)
            return False

    def verify_error_message(self, feature_file_name, screen_shot, Scenario):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_error_message(Scenario)
            self.logger.info("Successfully verified review icon")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_review_icon_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_review_icon_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_review_icon_verified_error.png"
            self.logger.error("Error while verifying review icon " + str(e))

    def do_click_apply_button(self, feature_file_name, screen_shot):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:

            x4a_bulk_order_upload.do_click_apply_button()
            self.logger.info("Successfully clicked on Apply button")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_status_selected_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_status_selected_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_file_reviewing_error.png"
            self.logger.error("Error while clicking Apply button %s", str(e))
            self.logger.exception(e)
            return False

    def verify_order_status(self, feature_file_name, screen_shot, status):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            x4a_bulk_order_upload.verify_order_status(status)
            self.logger.info("Successfully verified order status")
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                        + "_review_icon_verified_successfully.png")
            return True
        except Exception as e:
            self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name +
                                        "_review_icon_verified_error.png")
            screen_shot["path"] = self.screen_shot_path + "\\X4A\\error\\" + feature_file_name + \
                                  "_review_icon_verified_error.png"
            self.logger.error("Error while verifying order status " + str(e))

    def logout_x4a_url(self, feature_file_name):
        x4a_bulk_order_upload = X4ABulkOrderUploadPage(self.driver)
        try:
            if not x4a_bulk_order_upload.logout_x4a():
                self.logger.info("Failed to Logout X4A url")
                self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\error\\" + feature_file_name
                                            + "_logout_x4a_failed.png")
                return False
            else:
                self.logger.info("Successfully Logout X4A url")
                self.driver.save_screenshot(self.screen_shot_path + "\\X4A\\success\\" + feature_file_name
                                            + "_logout_x4a_successfully.png")
                return True
        except Exception as e:
            self.logger.error("Error while Logout the X4A url")
            self.logger.exception(e)
            return False
