import datetime
import json
import os

import pandas
import pandas as pd
from openpyxl import load_workbook, Workbook
from CommonUtilities.file_operations import logger
from CommonUtilities.logGeneration import LogGenerator
from CommonUtilities.readProperties import ReadConfig
from tests.test_script_constant import TestCaseID

# logger = LogGenerator.logGen()
"""This Module reads all test data and Writes output"""


# Returns total number of Rows when file is excel
def get_row_count(file, work_sheet_name):
    sheet = pd.read_excel(file, sheet_name=work_sheet_name)
    return len(sheet)


# Returns total number of Columns when file is excel
def get_column_count(file, work_sheet_name):
    sheet = pd.read_excel(file, sheet_name=work_sheet_name)
    return len(sheet.columns)


# Returns the value of a particular cell
def read_cell_data(file, row_number, column_number, work_sheet_name):
    """
    workbook = load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row_number, column_number).value
    """
    sheet = pd.read_excel(file, sheet_name=work_sheet_name)
    return sheet.iloc[row_number, column_number]


# Writes data in a single cell or a full row based on column_number input
def write_data_excel(file, data, row_number, column_number=None):
    # This method is not yet change for pandas as it is unused
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Output"
    if column_number is not None:
        sheet.cell(row_number, column_number, value=data)
    else:
        for column in range(len(data)):
            sheet.cell(row_number, column + 1, value=data[column])
    workbook.save(file)


def load_excel_to_dictionary(file, work_sheet_name):
    """
    test_data = []
    for row in range(getRowCount(file, sheet_name)-1):
        test_data_row = dict()
        for col in range(getColumnCount(file, sheet_name)):
            test_data_row.update({str(readCellData(file, 1, col+1, sheet_name)).strip(): str(readCellData(file, row+2, col+1, sheet_name)).strip()})
        test_data.append(test_data_row)
    return test_data
    """
    test_data = pd.read_excel(file, sheet_name=work_sheet_name)
    return test_data


def load_json_to_dictionary(file):
    my_json_file = open(file)
    test_data = json.load(my_json_file)
    my_json_file.close()
    return test_data


def read_cell_data_openpyxl(file, row, column, sheet_name):
    workbook = load_workbook(file)
    sheet = workbook[sheet_name]
    value = sheet.cell(row, column).value
    return value


def write_data_excel_openpyxl(file, data, sheet_name, row_number, column_number=None):
    workbook = load_workbook(file)
    sheet = workbook[sheet_name]
    if column_number is not None:
        sheet.cell(row_number, column_number, value=data)
    else:
        for column in range(len(data)):
            sheet.cell(row_number, column + 1, value=data[column])
    workbook.save(file)


def get_data_field_for_tc(file, sheet_name, column_number, test_case_id):
    data = ""
    for row in range(2, get_row_count(file, sheet_name) + 2):
        if read_cell_data_openpyxl(file, row, 1, sheet_name) == test_case_id:
            data = read_cell_data_openpyxl(file, row, column_number, sheet_name)
            break
    return data


""" Function update whole Test Data Sheet """


def update_test_data_sheet(file, sheet_name, test_case_id=None):
    date_time_obj = datetime.datetime.now()
    date_time_str = date_time_obj.strftime("%d%m%y_%H%M")
    for row in range(2, get_row_count(file, sheet_name) + 2):
        quote_name = company_name = read_cell_data_openpyxl(file, row, 3, sheet_name)[
                                    :9] + "_" + read_cell_data_openpyxl(
            file, row, 4, sheet_name) + "_" + read_cell_data_openpyxl(
            file, row, 1, sheet_name) + "_" + read_cell_data_openpyxl(
            file, row, 48, sheet_name) + "_" + date_time_str
        if test_case_id and test_case_id != read_cell_data_openpyxl(file, row, 1, sheet_name):
            continue
        if read_cell_data_openpyxl(file, row, 1, sheet_name) == TestCaseID.TS001:
            write_data_excel_openpyxl(file, quote_name, sheet_name, row, 3)
            write_data_excel_openpyxl(file, company_name, sheet_name, row, 22)
        elif read_cell_data_openpyxl(file, row, 1, sheet_name) == TestCaseID.TS022 or read_cell_data_openpyxl(file, row,
                                                                                                              1,
                                                                                                              sheet_name) == TestCaseID.TS026:
            write_data_excel_openpyxl(file, quote_name, sheet_name, row, 3)
            company_name = get_data_field_for_tc(file, sheet_name, 22, TestCaseID.TS001)
            if company_name:
                write_data_excel_openpyxl(file, company_name, sheet_name, row, 22)
            else:
                logger.info(
                    "TS001 testdata not present in sheet to copy company name so will use the existing data or can change manually")
        else:
            write_data_excel_openpyxl(file, quote_name, sheet_name, row, 3)
            write_data_excel_openpyxl(file, company_name, sheet_name, row, 22)
        if test_case_id:
            break


def update_test_data_sheet1(file, sheet_name, test_case_id=None):
    date_time_obj = datetime.datetime.now()
    date_time_str = date_time_obj.strftime("%d%m%y_%H%M")
    for row in range(2, get_row_count(file, sheet_name) + 2):
        quote_name = read_cell_data_openpyxl(file, row, 3, sheet_name)[
                     :9] + "_" + read_cell_data_openpyxl(
            file, row, 4, sheet_name) + "_" + read_cell_data_openpyxl(
            file, row, 1, sheet_name) + "_" + read_cell_data_openpyxl(
            file, row, 48, sheet_name) + "_" + date_time_str
        if test_case_id and test_case_id != read_cell_data_openpyxl(file, row, 1, sheet_name):
            continue
        if read_cell_data_openpyxl(file, row, 1, sheet_name) == TestCaseID.TS001:
            write_data_excel_openpyxl(file, quote_name, sheet_name, row, 3)
        elif read_cell_data_openpyxl(file, row, 1, sheet_name) == TestCaseID.TS022 or read_cell_data_openpyxl(file, row,
                                                                                                              1,
                                                                                                              sheet_name) == TestCaseID.TS026:
            write_data_excel_openpyxl(file, quote_name, sheet_name, row, 3)
            company_name = get_data_field_for_tc(file, sheet_name, 22, TestCaseID.TS001)
            if company_name:
                write_data_excel_openpyxl(file, company_name, sheet_name, row, 22)
            else:
                logger.info(
                    "TS001 testdata not present in sheet to copy company name so will use the existing data or can change manually")
        else:
            write_data_excel_openpyxl(file, quote_name, sheet_name, row, 3)

        if test_case_id:
            break


def get_marketplace(file, sheet_name):
    marketplaces_list = []
    for row in range(2, get_row_count(file, sheet_name) + 2):
        marketplace = read_cell_data_openpyxl(file, row, 48, sheet_name)
        marketplaces_list.append(marketplace.lower())
    distinct_marketplace = list(set(marketplaces_list))
    return distinct_marketplace


def get_data_from_tc_id_and_sku(file, sheet_name, column_number, test_case_id, sku):
    data = ""
    for row in range(2, get_row_count(file, sheet_name) + 2):
        if read_cell_data_openpyxl(file, row, 1, sheet_name) == test_case_id and read_cell_data_openpyxl(file, row, 2,
                                                                                                         sheet_name) == sku:
            data = read_cell_data_openpyxl(file, row, column_number, sheet_name)
            break
    return data


def get_test_case_data(path, worksheet_name):
    list = []
    test_data = load_excel_to_dictionary(path, worksheet_name)
    test_cases_list = test_data['Test cases'].tolist()
    for test_case in test_cases_list:
        controller_filtered_data = test_data.loc[
            test_data['Test cases'] == test_case]  # filter row with maching Test case ID
        print(controller_filtered_data)
        controller_filtered_data = controller_filtered_data.drop(
            columns=controller_filtered_data.columns[(controller_filtered_data == 'N').any()])
        print(controller_filtered_data)
        controller_filtered_data = controller_filtered_data.dropna(axis=1, how='all')  # remove empty columns
        print(controller_filtered_data)
        for order_index, test_data_order in controller_filtered_data.iterrows():
            print(test_data_order)
            original_dict = test_data_order.to_dict()
            print(original_dict)
            list.append(original_dict)
            print(list)
    return list


def get_marketplace_filtered_data(path, worksheet_name):
    test_data = load_excel_to_dictionary(path, worksheet_name)
    test_cases_list = test_data['TestCaseId'].tolist()
    marketplace_filtered_data = pd.read_excel(path)
    marketplace_filtered_data = marketplace_filtered_data[0:0]  # Creating a data frame and adding only headers to DF
    for test_case in test_cases_list:
        controller_filtered_data = test_data.loc[
            test_data['TestCaseId'] == test_case]  # filter row with matching Test case ID
        controller_filtered_data = controller_filtered_data.drop(
            columns=controller_filtered_data.columns[(controller_filtered_data == 'N').any()])
        controller_filtered_data = controller_filtered_data.dropna(axis=1, how='all')  # remove empty columns
        for name, values in controller_filtered_data.loc[:, controller_filtered_data.columns != 'TestCaseId'].items():
            data_sheet_name = "Data-" + name
            input_file_path = ReadConfig.get_test_data_file()
            marketplace_data = load_excel_to_dictionary(input_file_path, data_sheet_name)
            temp_dataframe = marketplace_data.loc[
                (marketplace_data['TestCaseId'] == test_case)]
            marketplace_filtered_data = marketplace_filtered_data.append(temp_dataframe, ignore_index=True)
    return marketplace_filtered_data


def get_erp_name(path, worksheet_name):
    erp_name = None
    try:
        region_data = load_excel_to_dictionary(path, worksheet_name)
        country_list = region_data['CountryCode'].tolist()
        logger.info(country_list)
        for country_code in country_list:
            marketplace = os.environ.get('CURRENT_MARKETPLACE')
            if country_code == marketplace:
                temp_dataframe = region_data.loc[region_data['CountryCode'] == marketplace]
                e_name = temp_dataframe['ERP'].tolist()
                erp_name = e_name[0]
                logger.info(f"ERP Name fetched successfully : {erp_name}. ")
                break
        return erp_name
    except Exception as e:
        logger.error(f"Exception occurred while trying to fetch the ERP name" + str(e))
        raise e


def get_company_code(path, worksheet_name):
    company_code = None
    try:
        region_data = load_excel_to_dictionary(path, worksheet_name)
        country_list = region_data['Country'].tolist()
        logger.info(country_list)
        for country_code in country_list:
            marketplace = os.environ.get('CURRENT_MARKETPLACE')
            if country_code == marketplace:
                temp_dataframe = region_data.loc[region_data['Country'] == marketplace]
                cmp_code = temp_dataframe['CompanyCode'].tolist()
                company_code = cmp_code[0]
                logger.info(f"Company Code fetched successfully: {company_code}. ")
                break
        return company_code
    except Exception as e:
        logger.error(f"Exception occurred while trying to fetch the company code" + str(e))
        raise e

def get_region_name(path, worksheet_name):
    region_name = None
    try:
        region_data = load_excel_to_dictionary(path, worksheet_name)
        country_list = region_data['CountryCode'].tolist()
        logger.info(country_list)
        for country_code in country_list:
            marketplace = os.environ.get('CURRENT_MARKETPLACE')
            if country_code == marketplace:
                temp_dataframe = region_data.loc[region_data['CountryCode'] == marketplace]
                r_name = temp_dataframe['Region'].tolist()
                region_name = r_name[0]
                logger.info(f"Region Name fetched successfully : {region_name}. ")
                break
        return region_name
    except Exception as e:
        logger.error(f"Exception occurred while trying to fetch the ERP name" + str(e))
        raise e


def fetch_test_case_filtered_data(path, worksheet_name ,test_case_id):
    test_data = load_excel_to_dictionary(path, worksheet_name)
    temp_dataframe = test_data.loc[(test_data['TestCaseId'] == test_case_id)]
    test_data_dictionary = temp_dataframe.to_dict('list')
    filtered_test_data = {}
    for key, value in test_data_dictionary.items():
        for val in value:
            filtered_test_data[key] = val
    return filtered_test_data
